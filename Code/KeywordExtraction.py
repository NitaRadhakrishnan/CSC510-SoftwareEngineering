# Libraries for text preprocessing
import pandas
import re
import nltk
import openpyxl as xl
import json
from scipy.sparse import coo_matrix
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import TfidfTransformer
from nltk.corpus import stopwords

nltk.download('stopwords')
from nltk.stem.porter import PorterStemmer
from nltk.tokenize import RegexpTokenizer

nltk.download('wordnet')
from nltk.stem.wordnet import WordNetLemmatizer

with open("data.json") as json_file:
    jsonData = json.load(json_file)


# Function for extracting keywords from user's message
def keywordExtraction(msg):
    text = msg

    ##Creating a list of stop words and adding custom stopwords
    stop_words = set(stopwords.words("english"))
    ##Creating a list of custom stopwords
    new_words = jsonData["newStopWords"]
    stop_words = stop_words.union(new_words)

    corpus = []

    # Convert to lowercase
    text = text.lower()

    # Convert to list from string
    text = text.split()

    # Stemming
    ps = PorterStemmer()
    # Lemmatisation
    lem = WordNetLemmatizer()
    text = [lem.lemmatize(word) for word in text if not word in stop_words]
    text = " ".join(text)
    corpus.append(text)
    # Text Preparation (Convert the words in the corpus to Tokens or vectors - Tokenization/Vectorization)
    # Creating a vector of word counts

    cv = CountVectorizer(min_df=0.8, stop_words=stop_words, max_features=10000, ngram_range=(1, 1))
    X = cv.fit_transform(corpus)
    list(cv.vocabulary_.keys())[:10]
    tfidf_transformer = TfidfTransformer(smooth_idf=True, use_idf=True)
    tfidf_transformer.fit(X)
    # get feature names
    feature_names = cv.get_feature_names()
    # fetch document for which keywords needs to be extracted
    for text1 in corpus:
        doc = text1
        # generate tf-idf for the given document
        tf_idf_vector = tfidf_transformer.transform(cv.transform([doc]))
    # sort the tf-idf vectors by descending order of scores
    sorted_items = sort_coo(tf_idf_vector.tocoo())
    # extract only the top n; n here is 10
    keywords = extract_topn_from_vector(feature_names, sorted_items, 5)
    # print(keyword_list)
    return (keywordlist(keywords))


# Function for sorting tf_idf in descending order
def sort_coo(coo_matrix):
    tuples = zip(coo_matrix.col, coo_matrix.data)
    return sorted(tuples, key=lambda x: (x[1], x[0]), reverse=True)


# Function to extract top 10 keywords based on the TF-IDF Scores
def extract_topn_from_vector(feature_names, sorted_items, topn=10):
    """get the feature names and tf-idf score of top n items"""
    # use only topn items from vector
    sorted_items = sorted_items[:topn]
    score_vals = []
    feature_vals = []

    # word index and corresponding tf-idf score
    for idx, score in sorted_items:
        # keep track of feature name and its corresponding score
        score_vals.append(round(score, 3))
        feature_vals.append(feature_names[idx])

    # create a tuples of feature,score
    # results = zip(feature_vals,score_vals)
    results = {}
    for idx in range(len(feature_vals)):
        results[feature_vals[idx]] = score_vals[idx]
    return results


def keywordlist(keywords):
    keyword_list = []
    for k in keywords:
        keyword_list.append(str(k))
    wb = xl.load_workbook(jsonData["xlsx_file"])
    sheet = wb['Sheet1']
    libInfo = {}
    listdict = []
    for row in range(2, sheet.max_row + 1):
        libInfo[sheet.cell(row, 1).value] = sheet.cell(row, 2).value
    for k in keyword_list:
        for r in (libInfo):
            if k in r:
                listdict.append({r: libInfo[r]})
    return (listdict)
